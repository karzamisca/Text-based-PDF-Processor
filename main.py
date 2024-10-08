import sys
import os
from PyQt6 import QtWidgets
from PyQt6.QtWidgets import QVBoxLayout, QLabel, QLineEdit, QPushButton, QFileDialog, QComboBox, QMessageBox
from pdf2docx import Converter
from docx import Document
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import re

# Function to normalize different quote and bracket characters
def normalize_text(text):
    # Normalize curly quotes to straight quotes
    text = re.sub(r'[“”]', '"', text)  # Replace both curly quotes with straight quotes
    # Normalize different types of brackets
    text = re.sub(r'[\uFF3D\u005D]', ']', text)  # Normalize full-width ] and standard ]
    return text

# Function to convert numbers into ordinal form (e.g., 1 to 1st, 2 to 2nd)
def ordinal(n):
    if 10 <= n % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
    return f"{n}{suffix}"

def pdf_to_word(pdf_file, word_file):
    # Convert PDF to Word using pdf2docx
    cv = Converter(pdf_file)
    cv.convert(word_file, start=0, end=None)
    cv.close()

def extract_paragraphs(word_file):
    # Open the Word document using python-docx
    doc = Document(word_file)
    paragraphs = [para.text for para in doc.paragraphs if para.text.strip() != ""]

    # Merge paragraphs split by page breaks and normalize the text
    merged_paragraphs = []
    temp_para = ""
    
    for para in paragraphs:
        # Normalize quotes and brackets in the paragraph
        para = normalize_text(para)

        if temp_para:
            temp_para += " " + para
        else:
            temp_para = para

        # Check if the paragraph ends with sentence-ending punctuation, including `]` and `"`
        if temp_para.endswith((".", "!", "?", ":", ";", "]", '"')):
            merged_paragraphs.append(temp_para)
            temp_para = ""

    if temp_para:
        merged_paragraphs.append(temp_para)

    # Now split each paragraph by line breaks
    final_paragraphs = []
    for para in merged_paragraphs:
        # Split paragraph by line breaks (\n or \r\n)
        lines = [line.strip() for line in re.split(r'\r?\n', para) if line.strip()]
        final_paragraphs.extend(lines)

    return final_paragraphs

def export_paragraphs_to_excel(paragraphs, excel_file):
    # Create a new Excel workbook and add data
    wb = Workbook()
    ws = wb.active
    ws.title = "Paragraphs"

    # Add column headers
    ws.append(["Ordinal", "Paragraph"])

    # Write each paragraph into a new row in the Excel file
    for idx, para in enumerate(paragraphs, start=1):  # Start from row 1 for ordinal numbers
        ws.append([ordinal(idx), para])
        cell = ws[f'B{idx+1}']  # Paragraphs are in column B, row idx+1 (header in row 1)
        
        # Set wrap text to true to show long paragraphs on multiple lines
        cell.alignment = Alignment(wrap_text=True)

    # Set column width to be wide enough for paragraphs (optional)
    ws.column_dimensions[get_column_letter(2)].width = 80  # Adjust width for column B (Paragraph)
    ws.column_dimensions[get_column_letter(1)].width = 15  # Adjust width for column A (Ordinal)
    
    # Auto-adjust row heights based on content
    for row in ws.iter_rows(min_row=2, max_row=len(paragraphs)+1, max_col=2):
        for cell in row:
            ws.row_dimensions[cell.row].height = 60  # Adjust row height (optional)

    # Save the Excel file
    wb.save(excel_file)

class PDFToExcelConverterApp(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PDF to Excel Converter")
        
        # Layout
        layout = QVBoxLayout()
        
        # Dropdown for file or folder selection
        self.file_or_folder_label = QLabel("Select Input Type:")
        self.file_or_folder_combo = QComboBox()
        self.file_or_folder_combo.addItems(["File", "Folder"])
        layout.addWidget(self.file_or_folder_label)
        layout.addWidget(self.file_or_folder_combo)

        # Input path selection
        self.input_path_label = QLabel("Input Path:")
        self.input_path_edit = QLineEdit()
        self.input_path_button = QPushButton("Browse")
        self.input_path_button.clicked.connect(self.browse_input)
        layout.addWidget(self.input_path_label)
        layout.addWidget(self.input_path_edit)
        layout.addWidget(self.input_path_button)

        # Output path selection
        self.output_path_label = QLabel("Output Path:")
        self.output_path_edit = QLineEdit()
        self.output_path_button = QPushButton("Browse")
        self.output_path_button.clicked.connect(self.browse_output)
        layout.addWidget(self.output_path_label)
        layout.addWidget(self.output_path_edit)
        layout.addWidget(self.output_path_button)

        # Convert button
        self.convert_button = QPushButton("Convert")
        self.convert_button.clicked.connect(self.convert_files)
        layout.addWidget(self.convert_button)

        self.setLayout(layout)

    def browse_input(self):
        if self.file_or_folder_combo.currentText() == "File":
            options = QFileDialog.Option()
            pdf_file, _ = QFileDialog.getOpenFileName(self, "Select PDF File", "", "PDF Files (*.pdf);;All Files (*)", options=options)
            if pdf_file:
                self.input_path_edit.setText(pdf_file)
        else:
            folder = QFileDialog.getExistingDirectory(self, "Select Folder")
            if folder:
                self.input_path_edit.setText(folder)

    def browse_output(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.output_path_edit.setText(folder)

    def convert_files(self):
        input_path = self.input_path_edit.text()
        output_path = self.output_path_edit.text()
        
        if not input_path or not output_path:
            QMessageBox.warning(self, "Input/Output Error", "Please select both input and output paths.")
            return

        if self.file_or_folder_combo.currentText() == "File":
            if not input_path.endswith(".pdf"):
                QMessageBox.warning(self, "File Error", "Selected file is not a PDF.")
                return
            
            pdf_files = [input_path]  # Single file case
        else:
            pdf_files = [os.path.join(input_path, f) for f in os.listdir(input_path) if f.endswith(".pdf")]

        # Process each PDF file
        for pdf_file in pdf_files:
            file_name = os.path.splitext(os.path.basename(pdf_file))[0]  # Get the base name without extension
            output_folder = os.path.join(output_path, file_name)  # Create output folder based on file name
            os.makedirs(output_folder, exist_ok=True)  # Create the folder if it doesn't exist
            
            word_file = os.path.join(output_folder, f"{file_name}.docx")  # Define Word file path
            excel_file = os.path.join(output_folder, f"{file_name}_paragraphs.xlsx")  # Define Excel file path

            pdf_to_word(pdf_file, word_file)
            paragraphs = extract_paragraphs(word_file)
            export_paragraphs_to_excel(paragraphs, excel_file)

        QMessageBox.information(self, "Success", "Conversion completed successfully!")

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = PDFToExcelConverterApp()
    window.resize(400, 200)
    window.show()
    sys.exit(app.exec())
